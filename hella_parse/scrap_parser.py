from os import listdir
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

scrap_types = ['141 zagon',
    '144 ponovni zagon (po čiščenju orodja)',
    '144 ponovni zagon (zaradi lakirnice)',
    '1 nezalito, nedolito',
    '2 posedeno',
    '3 prelito',
    '4 počeno, zvito, deformirano',
    '5 opraskano (robot, trak)',
    '6 onesnaženo s tujki',
    '7 žarki, plastika',
    '9 enojni žarek',
    '180 spojna linija',
    '12 mehurčki',
    '14 pike, pikasto (razno)',
    '15 meglica (mat, siva površina)',
    '23 nitke',
    '52 ris (orodje)',
    '172 U - zanka (pentlja, hakeljček)',
    '428 črna(e) pika(e)',
    '182 hladen brizg (gramofonska plošča)',
    '183 napake zaradi dodatnih operacij'
]

def get_scrap_types():
    return scrap_types

def parseCsvFiles(
               path="./odpad",
               show = False,
               showMissing = False):

    result = {}
    
    shift_offsets = { 3: 3, 2: 57, 1: 116 }
    product_start = 5
    product_width = 6
        
    files = listdir(path)
    for file in files:
        
        if not file.endswith('.xlsx'):
            continue
        
        workbook = load_workbook(path + file, True, True)
        for sheet in workbook.worksheets:
            title = sheet.title
            if not title.startswith('X'):
                continue
            
            print('Sheet: ' + title)
            
            date = sheet['A2'].value
            
            if date is None:
                print('Finished file!')
                break
            
            date_str = date.strftime('%d.%m.%Y')
            
            if date_str in result:
                print('Date: ' + date_str + ' is already in the result!!!')
            
            print('Date: ' + date_str)
            
            daily_report = {}    # daily_report[shift][product_name][scrap_type]
            
            for shift in shift_offsets:
                print('Shift: ' + str(shift))
                
                shift_offset = shift_offsets[shift]
                
                daily_report[shift] = {}
                
                product_n = 0
                while True:
                    product_col_n = product_start + product_n*product_width
                    product_col = get_column_letter(product_col_n)
                    
                    product_name = sheet[product_col + str(shift_offset)].value
                    
                    if product_name is not None and product_name.startswith('='):    # the value is linked to another cell
                        product_name = sheet[product_name[1:]].value
                    
                    if product_name is None or product_name == '' or product_name == '0':
                        break
                    
                    print('Product: ' + product_name)
                    
                    daily_report[shift][product_name] = {}
                    
                    for fault_n in range(len(scrap_types)):
                        scrap_type = scrap_types[fault_n]
                        left_col = product_col
                        right_col = get_column_letter(product_col_n + 3)
                        
                        scrap_left = sheet[left_col + str(shift_offset + 4 + fault_n)].value
                        scrap_right = sheet[right_col + str(shift_offset + 4 + fault_n)].value
                        
                        if scrap_left is None or scrap_left == '':
                            scrap_left = 0
                        else:
                            scrap_left = int(scrap_left)
                            
                        if scrap_right is None or scrap_right == '':
                            scrap_right = 0
                        else:
                            scrap_right = int(scrap_right)
                            
                        total_scrap = scrap_left + scrap_right
                        
                        daily_report[shift][product_name][scrap_type] = total_scrap

                    product_n += 1
            
            result[date_str] = daily_report
            
    return result

def doSum(lis):
    try:
        lis = [ int(i) if i and i != ' ' and i != 'xy' else 0 for i in lis]
    except:
        print(lis)
        raise
    return sum(lis)

def mySplit(string, seperator = ","):
    if string == "": return []
    res = []
    i = 0
    j = 0
    quoted = False
    for char in string:
        if char == '"':
            quoted = not quoted
        elif char == seperator and not quoted:
            app = string[i:j]
##            if app and app[-1] == '"': app = app[:-1]
##            if app and app[0] == '"': app = app[1:]
            res.append(app)
            i = j + 1 ## don't want to include the seperator
        j += 1
    res.append(string[i:])
    return res
            

if __name__ == "__main__":
    parseCsvFiles(show = True, showMissing = True)
















